"""Export shaping and do-not-contact enforcement."""

from __future__ import annotations

import csv
from datetime import date

import pytest
from openpyxl import load_workbook

from orynx.compliance.suppression import add_suppression
from orynx.db.models import Author, AuthorContact, Book, BookAuthor, BookSource, Lead, Source
from orynx.export.builder import COLUMNS, build_rows
from orynx.export.csv_export import write_csv
from orynx.export.xlsx_export import write_xlsx


@pytest.fixture
def populated(session):
    session.add(Source(id="fake", name="Fake", kind="hybrid", trust=0.7))
    author = Author(
        display_name="Amara Nwosu", normalized_name="amara nwosu",
        dedupe_key="nwosu:a", website="https://amara.example",
    )
    other = Author(
        display_name="Peter Blake", normalized_name="peter blake", dedupe_key="blake:p"
    )
    book = Book(
        title="The Quiet Harbour", isbn13="9781234567897", publisher="Koehler Books",
        published_on=date(2026, 3, 15), published_year=2026, ratings_count=4,
        categories=["Fiction"], dedupe_key="isbn:9781234567897",
    )
    session.add_all([author, other, book])
    session.flush()
    session.add_all(
        [
            AuthorContact(author_id=author.id, kind="email", value="amara@amara.example",
                          source_id="fake", source_url="https://amara.example"),
            AuthorContact(author_id=author.id, kind="twitter", value="https://x.com/amara",
                          source_id="fake", source_url="https://amara.example"),
            BookAuthor(book_id=book.id, author_id=author.id, position=0),
            BookAuthor(book_id=book.id, author_id=other.id, position=1),
            BookSource(book_id=book.id, source_id="fake", external_id="1",
                       url="https://press.test/a"),
            Lead(author_id=author.id, book_id=book.id, score=88.0, tier="A",
                 reasons=[{"signal": "recency", "points": 28.0}]),
            Lead(author_id=other.id, book_id=book.id, score=41.0, tier="C", reasons=[]),
        ]
    )
    session.commit()
    return session


def test_rows_carry_contacts_and_provenance(populated):
    rows, suppressed = build_rows(populated)
    assert suppressed == 0
    top = rows[0]
    assert top.author_name == "Amara Nwosu"
    assert top.author_emails == "amara@amara.example"
    assert "twitter=" in top.author_socials
    assert top.sources == "fake"
    assert top.source_urls == "https://press.test/a"
    assert top.isbn13 == "9781234567897"
    assert "recency" in top.top_signals


def test_rows_are_ordered_by_score(populated):
    rows, _ = build_rows(populated)
    assert [r.score for r in rows] == sorted((r.score for r in rows), reverse=True)


def test_min_score_and_tier_filters(populated):
    assert len(build_rows(populated, min_score=50)[0]) == 1
    assert len(build_rows(populated, tiers=["A"])[0]) == 1
    assert len(build_rows(populated, tiers=["A", "C"])[0]) == 2


def test_require_contact_drops_uncontactable_leads(populated):
    rows, _ = build_rows(populated, require_contact=True)
    assert [r.author_name for r in rows] == ["Amara Nwosu"]


def test_suppressed_email_is_withheld(populated):
    add_suppression(populated, "email", "amara@amara.example", "opted out")
    populated.commit()
    rows, suppressed = build_rows(populated)
    assert suppressed == 1
    assert all(r.author_name != "Amara Nwosu" for r in rows)


def test_suppressed_domain_is_withheld(populated):
    add_suppression(populated, "domain", "amara.example")
    populated.commit()
    _, suppressed = build_rows(populated)
    assert suppressed == 1


def test_suppressed_author_name_is_withheld(populated):
    add_suppression(populated, "author_name", "peter blake")
    populated.commit()
    rows, suppressed = build_rows(populated)
    assert suppressed == 1
    assert all(r.author_name != "Peter Blake" for r in rows)


def test_suppression_can_be_bypassed_only_explicitly(populated):
    add_suppression(populated, "email", "amara@amara.example")
    populated.commit()
    rows, suppressed = build_rows(populated, apply_suppression=False)
    assert suppressed == 0
    assert len(rows) == 2


def test_csv_round_trips_every_column(populated, tmp_path):
    rows, _ = build_rows(populated)
    path = write_csv(rows, tmp_path / "leads.csv")
    with path.open(encoding="utf-8-sig") as handle:
        parsed = list(csv.DictReader(handle))
    assert list(parsed[0]) == COLUMNS
    assert parsed[0]["author_name"] == "Amara Nwosu"
    assert len(parsed) == 2


def test_xlsx_has_a_leads_sheet_and_a_summary(populated, tmp_path):
    rows, suppressed = build_rows(populated)
    path = write_xlsx(rows, tmp_path / "leads.xlsx", suppressed=suppressed)
    workbook = load_workbook(path)
    assert workbook.sheetnames == ["Leads", "Summary"]

    sheet = workbook["Leads"]
    assert [c.value for c in sheet[1]] == COLUMNS
    assert sheet.max_row == 3  # header plus two leads
    assert sheet.freeze_panes == "A2"
    # Score must land as a number, not text, so the sheet sorts on it correctly.
    # openpyxl narrows a whole float to int, so accept either.
    score_cell = sheet.cell(row=2, column=COLUMNS.index("score") + 1).value
    assert isinstance(score_cell, (int, float)) and not isinstance(score_cell, bool)

    summary = {r[0]: r[1] for r in workbook["Summary"].iter_rows(values_only=True) if r[0]}
    assert summary["Leads exported"] == 2
    assert summary["With an email"] == 1


def test_export_of_an_empty_database_is_not_an_error(session, tmp_path):
    rows, _ = build_rows(session)
    assert rows == []
    path = write_csv(rows, tmp_path / "empty.csv")
    assert path.exists()
