"""Excel output: a formatted lead sheet plus a summary of how it was built."""

from __future__ import annotations

from collections import Counter
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from orynx.export.builder import COLUMNS, LeadRow

HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TIER_FILLS = {
    "A": PatternFill("solid", fgColor="D1FAE5"),
    "B": PatternFill("solid", fgColor="FEF3C7"),
    "C": PatternFill("solid", fgColor="F3F4F6"),
    "D": PatternFill("solid", fgColor="FEE2E2"),
}
NUMERIC_COLUMNS = {"score", "page_count", "ratings_count", "average_rating", "published_year"}
WIDE_COLUMNS = {"book_title": 40, "author_name": 24, "publisher": 24, "categories": 30,
                "source_urls": 40, "top_signals": 34, "author_emails": 28}


def write_xlsx(rows: list[LeadRow], path: Path, *, suppressed: int = 0) -> Path:
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Leads"

    sheet.append(COLUMNS)
    for index, column in enumerate(COLUMNS, start=1):
        cell = sheet.cell(row=1, column=index)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center")
        sheet.column_dimensions[get_column_letter(index)].width = WIDE_COLUMNS.get(column, 16)

    for row in rows:
        data = row.as_dict()
        values = []
        for column in COLUMNS:
            value = data.get(column, "")
            if column in NUMERIC_COLUMNS and value not in ("", None):
                try:
                    value = float(value)
                except (TypeError, ValueError):
                    pass
            values.append(value)
        sheet.append(values)
        fill = TIER_FILLS.get(row.tier)
        if fill:
            sheet.cell(row=sheet.max_row, column=COLUMNS.index("tier") + 1).fill = fill

    sheet.freeze_panes = "A2"
    if rows:
        sheet.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{sheet.max_row}"

    summary = workbook.create_sheet("Summary")
    tier_counts = Counter(row.tier for row in rows)
    source_counts: Counter[str] = Counter()
    for row in rows:
        for source in row.sources.split("; "):
            if source:
                source_counts[source] += 1

    summary.append(["Metric", "Value"])
    summary["A1"].font = HEADER_FONT
    summary["B1"].font = HEADER_FONT
    summary["A1"].fill = HEADER_FILL
    summary["B1"].fill = HEADER_FILL
    summary.append(["Leads exported", len(rows)])
    summary.append(["Suppressed (do-not-contact)", suppressed])
    summary.append(["With an email", sum(1 for r in rows if r.author_emails)])
    reachable = sum(
        1 for r in rows if r.author_emails or r.author_website or r.author_socials
    )
    summary.append(["With any contact point", reachable])
    summary.append([])
    summary.append(["Tier", "Count"])
    for tier in ("A", "B", "C", "D"):
        summary.append([tier, tier_counts.get(tier, 0)])
    summary.append([])
    summary.append(["Source", "Leads"])
    for source, count in source_counts.most_common():
        summary.append([source, count])
    summary.column_dimensions["A"].width = 32
    summary.column_dimensions["B"].width = 14

    workbook.save(path)
    return path
